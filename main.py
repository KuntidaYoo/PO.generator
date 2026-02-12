import os
import re
import math
import datetime
from typing import List, Dict, Optional, Tuple
from copy import copy as _copy
from io import BytesIO

import numpy as np
import pandas as pd
import openpyxl
from PIL import Image as PILImage

from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Border, Side, PatternFill
from openpyxl.utils.cell import coordinate_from_string, column_index_from_string
from openpyxl.utils.units import pixels_to_EMU
from openpyxl.drawing.spreadsheet_drawing import OneCellAnchor, AnchorMarker
from openpyxl.drawing.xdr import XDRPositiveSize2D


# =========================
# CONFIG (template/layout)
# =========================
EXPRESS_SHEET = 0

TEMPLATE_SHEET_NAME = "page1"
HEADER_ROW = 8
ITEM_START_ROW = 9
TEMPLATE_ITEM_ROW = 9
TEMPLATE_TOTAL_START_ROW = 14
TEMPLATE_TOTAL_END_ROW = 16

PO_OUTPUT_FOLDER = "output_PO"

IMAGE_WIDTH_BOOST = 1.20
IMAGE_PADDING_PX = 2

HIGHLIGHT_BELOW_MIN = PatternFill(fill_type="solid", start_color="FFF2CC", end_color="FFF2CC")


# =========================
# THAI DATE
# =========================
THAI_MONTHS = {
    "ม.ค.": 1, "ก.พ.": 2, "มี.ค.": 3, "เม.ย.": 4,
    "พ.ค.": 5, "มิ.ย.": 6, "ก.ค.": 7, "ส.ค.": 8,
    "ก.ย.": 9, "ต.ค.": 10, "พ.ย.": 11, "ธ.ค.": 12
}


def thai_to_date(day: int, thai_month: str, thai_year: int) -> datetime.date:
    """Convert Thai BE date tokens into Gregorian date."""
    year = thai_year - 543
    thai_month = re.sub(r"\s+", "", thai_month)
    month = THAI_MONTHS.get(thai_month)
    if month is None:
        raise ValueError(f"Unknown Thai month token: {repr(thai_month)}")
    return datetime.date(year, month, day)


def last_day_of_month(year: int, month: int) -> int:
    """Return last day number of a given month."""
    if month == 12:
        next_first = datetime.date(year + 1, 1, 1)
    else:
        next_first = datetime.date(year, month + 1, 1)
    return (next_first - datetime.timedelta(days=1)).day


def add_months(dt: datetime.date, n: int) -> datetime.date:
    """Add N months to a date while keeping a valid day."""
    month = dt.month - 1 + n
    year = dt.year + month // 12
    month = month % 12 + 1
    day = min(dt.day, last_day_of_month(year, month))
    return datetime.date(year, month, day)


def calc_days_and_months(d1: datetime.date, d2: datetime.date) -> Tuple[int, int]:
    """
    Compute (days, months) where leftover >= 15 days counts as an extra month.
    """
    if d2 < d1:
        return 0, 0
    days = (d2 - d1).days
    months = 0
    cur = d1
    while True:
        nxt = add_months(cur, 1)
        if nxt <= d2:
            months += 1
            cur = nxt
        else:
            break
    leftover = (d2 - cur).days
    if leftover >= 15:
        months += 1
    return days, months


def parse_date_range_from_header(df_raw: pd.DataFrame) -> Dict[str, object]:
    """
    Extract Thai date range text like:
      'วันที่จาก 1 ม.ค. 2569 ถึง 31 ม.ค. 2569'
    and convert to months by the 15-day rule.
    """
    pattern = r"วันที่จาก\s+(\d+)\s+(\S+)\s+(\d+)\s+ถึง\s+(\d+)\s+(\S+)\s+(\d+)"
    col0 = df_raw.iloc[:, 0].astype(str)
    for val in col0:
        if "วันที่จาก" not in str(val):
            continue
        text = str(val).replace("\xa0", " ")
        text = re.sub(r"\s+", " ", text).strip()
        m = re.search(pattern, text)
        if not m:
            continue
        d1, m1, y1, d2, m2, y2 = m.groups()
        start_date = thai_to_date(int(d1), m1, int(y1))
        end_date = thai_to_date(int(d2), m2, int(y2))
        days, months = calc_days_and_months(start_date, end_date)
        return {"raw_line": text, "start_date": start_date, "end_date": end_date, "days": days, "months": months}
    return {"raw_line": "", "start_date": None, "end_date": None, "days": 0, "months": 1}


# =========================
# SMALL UTILS
# =========================
def round_half_up(x: float) -> int:
    """Round half up (0.5 -> 1)."""
    return int(math.floor(x + 0.5))


def norm_text(v) -> str:
    """Normalize cell text."""
    if v is None:
        return ""
    return str(v).replace("\xa0", " ").strip()


def clean_cell(val) -> str:
    """Clean a raw cell value into single-space text."""
    if pd.isna(val):
        return ""
    s = str(val).replace("\xa0", " ")
    s = re.sub(r"\s+", " ", s).strip()
    return s


def fix_split_numbers_in_line(line: str) -> str:
    """
    Fix patterns where a number is split into two tokens (rare Excel formatting artifact).
    Example: '1' '234.00' -> '1234.00'
    """
    tokens = line.split()
    i = 0
    while i < len(tokens) - 1:
        cur = tokens[i]
        nxt = tokens[i + 1]
        if re.fullmatch(r"\d{1,2}", cur) and re.fullmatch(r"\d{3,}(?:\.\d+)?\"?", nxt):
            tokens[i] = cur + nxt
            del tokens[i + 1]
        else:
            i += 1
    return " ".join(tokens)


def row_to_merged_line(row: pd.Series) -> str:
    """
    Merge all non-empty cells for robust parsing of buyer/product text.
    (We do NOT rely on merging for yuan; yuan is found by scanning row cells.)
    """
    cells = [clean_cell(v) for v in row.tolist()]
    cells = [c for c in cells if c]
    raw = " ".join(cells).strip()
    return fix_split_numbers_in_line(raw)


def looks_like_buyer_code(token: str) -> bool:
    """Buyer code is 5 alphanumeric characters."""
    return bool(re.fullmatch(r"[0-9A-Za-z]{5}", token or ""))


def is_header_or_separator(line: str) -> bool:
    """Skip header-like rows."""
    s = (line or "").strip()
    if not s:
        return True
    if "BUYER" in s:
        return True
    if "วันที่จาก" in s:
        return True
    if re.search(r"-{5,}", s):
        return True
    return False


# =========================
# PRODUCT SPLIT
# =========================
_DASH_TRANS = str.maketrans({
    "‐": "-", "-": "-", "‒": "-", "–": "-", "—": "-", "−": "-",
})


def split_product_field(s: str) -> Tuple[str, str]:
    """
    Split merged product string into (product_code, description).

    Supports:
      - doc token like 01-15-0730-D3-5 (ignore)
      - codes like NR123-60, DGS-2318, MC401-18, BM-150, NRW, etc
      - NoBM-150 => BM-150
    """
    if not isinstance(s, str):
        return "", ""
    s = s.strip()
    if not s:
        return "", ""

    s = s.translate(_DASH_TRANS)
    parts = s.split(maxsplit=1)
    if len(parts) < 2:
        return "", ""
    rest = parts[1].strip()
    if not rest:
        return "", ""

    def is_doc_token(t: str) -> bool:
        t = t.strip().translate(_DASH_TRANS)
        return bool(re.fullmatch(r"\d{2}-\d{2}-\d{3,6}(?:-[A-Za-z0-9]+)*", t))

    def is_product_code(t: str) -> bool:
        t = t.strip().translate(_DASH_TRANS)
        t2 = re.sub(r'^[Nn][Oo](?=[A-Za-z0-9])', '', t).strip()
        if not t2:
            return False
        if re.fullmatch(r"[A-Za-z]{1,6}[A-Za-z0-9]*-[A-Za-z0-9]+", t2):
            return True
        if re.fullmatch(r"[A-Za-z]{1,6}\d+[A-Za-z0-9]*", t2):
            return True
        if re.fullmatch(r"[A-Za-z]{2,5}", t2):  # e.g. NRW
            return True
        return False

    m_th = re.search(r"[\u0E00-\u0E7F]", rest)
    if m_th:
        th_pos = m_th.start()
        pre_th = rest[:th_pos].strip()
        tail_th = rest[th_pos:].strip()
        tokens = pre_th.split()
        if not tokens:
            return "", tail_th

        while tokens and is_doc_token(tokens[0]):
            tokens = tokens[1:]
        if not tokens:
            return "", tail_th

        code_idx = None
        for i, t in enumerate(tokens):
            if is_product_code(t):
                code_idx = i
                break
        if code_idx is None:
            code_idx = 0

        code_raw = tokens[code_idx].translate(_DASH_TRANS)
        code = re.sub(r'^[Nn][Oo](?=[A-Za-z0-9])', '', code_raw).strip()

        desc_pre = " ".join(tokens[code_idx + 1:]).strip()
        desc = (desc_pre + " " + tail_th).strip() if desc_pre else tail_th
        return code, desc

    tokens = rest.split()
    while tokens and is_doc_token(tokens[0]):
        tokens = tokens[1:]
    if not tokens:
        return "", ""

    code_raw = tokens[0].translate(_DASH_TRANS)
    code = re.sub(r'^[Nn][Oo](?=[A-Za-z0-9])', '', code_raw).strip()
    desc = " ".join(tokens[1:]).strip()
    return code, desc


# =========================
# MONEY BLOCK + 5/6 NUMBERS
# =========================
_MONEY_2DP_RE = re.compile(r"-?\d+(?:,\d{3})*\.\d{2}")
_YUAN_RE = re.compile(r"^[Yy]?\s*(-?\d+(?:\.\d+)?)\s*$")


def _money_tokens_2dp_in_text(s: str) -> List[float]:
    out: List[float] = []
    if not s:
        return out
    for raw in _MONEY_2DP_RE.findall(s):
        try:
            out.append(float(raw.replace(",", "")))
        except Exception:
            pass
    return out


def extract_money_2dp_numbers_from_row(row: pd.Series, scan_cols: int = 25) -> List[float]:
    """
    Extract 2dp money numbers from the first `scan_cols` columns (left area).
    Keeps the order left->right.
    """
    out: List[float] = []
    for i in range(min(scan_cols, len(row))):
        v = row.iloc[i]
        if v is None:
            continue
        s = str(v).replace("\xa0", " ").strip()
        if not s:
            continue
        out.extend(_money_tokens_2dp_in_text(s))
    return out


def find_money_block_cell(row: pd.Series, min_tokens: int = 3) -> int:
    """
    Return money_block_cell_index (0-based) where a single cell contains >= min_tokens money 2dp tokens.
    If none found, return -1.
    """
    for i in range(len(row)):
        v = row.iloc[i]
        if v is None:
            continue
        s = str(v).replace("\xa0", " ").strip()
        if not s:
            continue
        if len(_MONEY_2DP_RE.findall(s)) >= min_tokens:
            return i
    return -1


def parse_yuan_value(v) -> Optional[float]:
    """
    Accept:
      - 7.98
      - Y7.98 / Y 7.98
    Reject:
      - % values
      - junk text
      - 0 / 0.00
    """
    if v is None:
        return None
    s = str(v).replace("\xa0", " ").strip()
    if not s:
        return None
    if "%" in s:
        return None

    m = _YUAN_RE.match(s)
    if not m:
        return None
    try:
        val = float(m.group(1))
    except Exception:
        return None
    if abs(val) < 1e-12:
        return None
    return val


def extract_yuan_after_money_block(row: pd.Series, lookahead_cells: int = 20) -> Optional[float]:
    """
    FINAL RULE YOU CONFIRMED:
      - locate money-block cell (contains many 2dp numbers)
      - scan next 15-20 cells to the right
      - the first valid numeric cell there is yuan
      - if none -> None
    """
    money_idx = find_money_block_cell(row, min_tokens=3)
    if money_idx < 0:
        return None

    start = money_idx + 1
    end = min(len(row), money_idx + 1 + int(lookahead_cells))

    for j in range(start, end):
        v = row.iloc[j]
        if v is None:
            continue
        s = str(v).replace("\xa0", " ").strip()
        if not s:
            continue
        y = parse_yuan_value(s)
        if y is not None:
            return y
    return None


# =========================
# PARSE ONE LINE -> FIELDS
# =========================
def parse_line_to_fields(row: pd.Series, merged_line: str) -> Optional[Dict[str, object]]:
    """
    Parse an Express row into:
      buyer, barcode(optional), สินค้า(product_str), ยอดขาย, สินค้าคงเหลือ, ON_ORDER, หยวน
    The 5/6-number block is extracted from row (2dp tokens).
    Yuan is extracted by money-block anchor + lookahead scan.
    """
    m = re.match(r"\s*([0-9A-Za-z]{5})\b(.*)", merged_line or "")
    if not m:
        return None

    buyer = m.group(1).strip().upper()
    rest = m.group(2).strip()
    if not rest:
        return None

    tokens = rest.split()
    if not tokens:
        return None

    barcode = ""
    idx = 0
    if idx < len(tokens) and re.fullmatch(r"\d+", tokens[idx] or ""):
        barcode = tokens[idx]
        idx += 1

    product_str = " ".join(tokens[idx:]).strip()
    if not product_str:
        return None

    nums = extract_money_2dp_numbers_from_row(row, scan_cols=25)
    if len(nums) < 5:
        return None

    block = nums[-6:] if len(nums) >= 6 else nums[-5:]

    # your old mapping:
    # if 6 numbers: sale=block[2], stock=block[3], on_order=block[5]
    # if 5 numbers: sale=0, stock=block[2], on_order=block[4]
    if len(block) == 6:
        sale = float(block[2])
        stock = float(block[3])
        on_order = float(block[5])
    else:
        sale = 0.0
        stock = float(block[2])
        on_order = float(block[4])

    yuan_val = extract_yuan_after_money_block(row, lookahead_cells=20)

    return {
        "buyer": buyer,
        "barcode": barcode,
        "สินค้า": product_str,
        "ยอดขาย": sale,
        "สินค้าคงเหลือ": stock,
        "ON_ORDER": on_order,
        "หยวน": yuan_val,
    }


def parse_express_file(path: str, source_label: str) -> Tuple[pd.DataFrame, Dict[str, object]]:
    """
    Parse Express export into DataFrame:
      buyer, barcode, รหัสสินค้า, รายละเอียดสินค้า, ยอดขาย, สินค้าคงเหลือ, ON_ORDER, หยวน
    """
    df_raw = pd.read_excel(path, sheet_name=EXPRESS_SHEET, header=None, dtype=str)
    date_info = parse_date_range_from_header(df_raw)

    rows = []
    for idx, row in df_raw.iterrows():
        merged = row_to_merged_line(row)
        if not merged:
            continue
        if is_header_or_separator(merged):
            continue

        first_token = merged.split(maxsplit=1)[0] if merged.split() else ""
        if not looks_like_buyer_code(first_token):
            continue

        fields = parse_line_to_fields(row, merged)
        if fields is None:
            continue

        fields["source"] = source_label
        fields["src_row"] = int(idx) + 1
        fields["src_file"] = os.path.basename(path)
        rows.append(fields)

    df = pd.DataFrame(rows)
    if not df.empty:
        df["buyer"] = df["buyer"].astype(str).str.replace("\u00A0", " ", regex=False).str.strip().str.upper()
        df[["รหัสสินค้า", "รายละเอียดสินค้า"]] = df["สินค้า"].apply(lambda x: pd.Series(split_product_field(x)))
        df.drop(columns=["สินค้า"], inplace=True)

    return df, date_info


# =========================
# COMBINE + AGG
# =========================
def _agg_one(df: pd.DataFrame, label: str) -> pd.DataFrame:
    """Aggregate per (buyer, รหัสสินค้า)."""
    key_cols = ["buyer", "รหัสสินค้า"]
    if df.empty:
        return pd.DataFrame(columns=key_cols + [
            f"ยอดขาย_{label}", f"STOCK_{label}", f"ON_ORDER_{label}", f"หยวน_{label}",
            "barcode", "รายละเอียดสินค้า"
        ])

    g = df.groupby(key_cols, as_index=False).agg({
        "ยอดขาย": "sum",
        "สินค้าคงเหลือ": "sum",
        "ON_ORDER": "sum",
        "หยวน": "first",
        "barcode": "first",
        "รายละเอียดสินค้า": "first",
    })
    return g.rename(columns={
        "ยอดขาย": f"ยอดขาย_{label}",
        "สินค้าคงเหลือ": f"STOCK_{label}",
        "ON_ORDER": f"ON_ORDER_{label}",
        "หยวน": f"หยวน_{label}",
    })


def build_combined_all(df_asia: pd.DataFrame, df_green: pd.DataFrame, months: int, min_factor: int, max_factor: int) -> pd.DataFrame:
    """
    Build combined dataset for all buyers (no vendor filtering).
    """
    g_asia = _agg_one(df_asia, "ASIA")
    g_green = _agg_one(df_green, "GREEN")

    combined = pd.merge(g_asia, g_green, on=["buyer", "รหัสสินค้า"], how="outer", suffixes=("", "_dup"))

    def coalesce(a, b):
        return a if pd.notna(a) and a != "" else b

    combined["barcode"] = [coalesce(a, b) for a, b in zip(
        combined.get("barcode_x", [None] * len(combined)),
        combined.get("barcode_y", [None] * len(combined)),
    )]
    combined["รายละเอียดสินค้า"] = [coalesce(a, b) for a, b in zip(
        combined.get("รายละเอียดสินค้า_x", [None] * len(combined)),
        combined.get("รายละเอียดสินค้า_y", [None] * len(combined)),
    )]

    for col in ["barcode_x", "barcode_y", "รายละเอียดสินค้า_x", "รายละเอียดสินค้า_y"]:
        if col in combined.columns:
            combined.drop(columns=[col], inplace=True)

    for col in ["ยอดขาย_ASIA", "STOCK_ASIA", "ON_ORDER_ASIA",
                "ยอดขาย_GREEN", "STOCK_GREEN", "ON_ORDER_GREEN"]:
        if col not in combined.columns:
            combined[col] = 0.0
        else:
            combined[col] = combined[col].fillna(0.0)

    combined["ยอดขาย_TOTAL"] = combined["ยอดขาย_ASIA"] + combined["ยอดขาย_GREEN"]
    combined["ON_ORDER_TOTAL"] = combined["ON_ORDER_ASIA"] + combined["ON_ORDER_GREEN"]

    def pick_yuan(row):
        yG = row.get("หยวน_GREEN", np.nan)
        yA = row.get("หยวน_ASIA", np.nan)
        if pd.notna(yG):
            try:
                return float(yG)
            except Exception:
                return np.nan
        if pd.notna(yA):
            try:
                return float(yA)
            except Exception:
                return np.nan
        return np.nan

    combined["หยวน"] = combined.apply(pick_yuan, axis=1)

    if months <= 0:
        months = 1

    combined["USE_MONTH"] = combined["ยอดขาย_TOTAL"].apply(lambda v: round_half_up(v / months) if v > 0 else 0)
    combined["TOTAL_QTY_NUM"] = combined["STOCK_ASIA"] + combined["STOCK_GREEN"] + combined["ON_ORDER_TOTAL"]
    combined["MIN_NUM"] = combined["USE_MONTH"] * int(min_factor)
    combined["MAX_NUM"] = combined["USE_MONTH"] * int(max_factor)

    return combined


# =========================
# VENDOR INFO
# =========================
def load_vendor_map(path: str) -> dict:
    """Read vendor info file: column 0=code, 1=name, 2=address."""
    if not os.path.exists(path):
        return {}
    df = pd.read_excel(path, header=None)
    out = {}
    for _, r in df.iterrows():
        code = str(r.iloc[0]).strip() if not pd.isna(r.iloc[0]) else ""
        if not code:
            continue
        out[str(code).strip().upper()] = {
            "name": "" if pd.isna(r.iloc[1]) else str(r.iloc[1]).strip(),
            "address": "" if pd.isna(r.iloc[2]) else str(r.iloc[2]).strip(),
        }
    return out


# =========================
# CATALOG (multi-sheet per vendor)
# =========================
def build_catalog_map(catalog_path: str, vendor_code: str) -> dict:
    """
    Read catalog workbook where each vendor has its own sheet.
    Column mapping by Excel position:
      A=item no, B=picture, C=desc, D=brand, E=material, F=weight, G=qty/carton
    """
    wb = openpyxl.load_workbook(catalog_path)
    want = str(vendor_code).strip().upper()
    norm_map = {str(n).strip().upper(): n for n in wb.sheetnames}

    if want in norm_map:
        ws = wb[norm_map[want]]
    else:
        ws = None
        for k_norm, original in norm_map.items():
            if want in k_norm or k_norm in want:
                ws = wb[original]
                break
        if ws is None:
            raise RuntimeError(f"Catalog workbook has no sheet for vendor '{want}'. Available: {wb.sheetnames}")

    COL_ITEM_NO = 1
    COL_PIC = 2
    COL_DESC = 3
    COL_BRAND = 4
    COL_MAT = 5
    COL_WEIGHT = 6
    COL_QTYCT = 7
    HEADER_ROW_LOCAL = 1

    img_at = {}
    for img in ws._images:
        try:
            r = img.anchor._from.row + 1
            c = img.anchor._from.col + 1
            img_at[(r, c)] = img._data()
        except Exception:
            pass

    catalog = {}
    for r in range(HEADER_ROW_LOCAL + 1, ws.max_row + 1):
        item_no = ws.cell(r, COL_ITEM_NO).value
        if not item_no:
            continue
        item_no = str(item_no).strip()
        catalog[item_no] = {
            "goods_desc": ws.cell(r, COL_DESC).value,
            "brand": ws.cell(r, COL_BRAND).value,
            "material": ws.cell(r, COL_MAT).value,
            "weight": ws.cell(r, COL_WEIGHT).value,
            "qty_per_carton": ws.cell(r, COL_QTYCT).value,
            "img_bytes": img_at.get((r, COL_PIC)),
        }
    return catalog


# =========================
# EXCEL IMAGE + STYLE HELPERS
# =========================
def _excel_colwidth_to_pixels(width):
    """Convert Excel col width to pixels."""
    if width is None:
        width = 8.43
    return int(width * 7 + 5)


def _excel_rowheight_to_pixels(height_pts):
    """Convert Excel row height to pixels."""
    if height_pts is None:
        height_pts = 15
    return int(height_pts * 96 / 72)


def _get_cell_rect_pixels(ws, col_letter, row_num):
    """Return (w_px, h_px) including merged cell extents."""
    col_w = _excel_colwidth_to_pixels(ws.column_dimensions[col_letter].width)
    row_h = _excel_rowheight_to_pixels(ws.row_dimensions[row_num].height)

    for mr in ws.merged_cells.ranges:
        if mr.min_col <= column_index_from_string(col_letter) <= mr.max_col and mr.min_row <= row_num <= mr.max_row:
            total_w = 0
            for c in range(mr.min_col, mr.max_col + 1):
                letter = get_column_letter(c)
                total_w += _excel_colwidth_to_pixels(ws.column_dimensions[letter].width)
            total_h = 0
            for rr in range(mr.min_row, mr.max_row + 1):
                total_h += _excel_rowheight_to_pixels(ws.row_dimensions[rr].height)
            return total_w, total_h

    return col_w, row_h


def add_image_to_cell(ws, cell_addr: str, img_bytes: bytes,
                      width_boost: float = IMAGE_WIDTH_BOOST,
                      padding_px: int = IMAGE_PADDING_PX):
    """Place an image into an Excel cell with center alignment."""
    if not img_bytes:
        return

    col_letter, row_num = coordinate_from_string(cell_addr)
    row_num = int(row_num)

    cell_w_px, cell_h_px = _get_cell_rect_pixels(ws, col_letter, row_num)

    max_w = max(1, int((cell_w_px - padding_px) * width_boost))
    max_w = min(max_w, cell_w_px - padding_px)
    max_h = max(1, cell_h_px - padding_px)

    pil = PILImage.open(BytesIO(img_bytes)).convert("RGBA")
    w, h = pil.size
    scale = min(max_w / w, max_h / h, 1.0)
    new_w = max(1, int(w * scale))
    new_h = max(1, int(h * scale))
    pil = pil.resize((new_w, new_h))

    bio = BytesIO()
    pil.save(bio, format="PNG")
    bio.seek(0)

    img = XLImage(bio)
    img.width = new_w
    img.height = new_h

    if not hasattr(ws, "_img_buffers"):
        ws._img_buffers = []
    ws._img_buffers.append(bio)

    col_idx0 = column_index_from_string(col_letter) - 1
    row_idx0 = row_num - 1
    x_off_px = max(0, int((cell_w_px - new_w) / 2))
    y_off_px = max(0, int((cell_h_px - new_h) / 2))

    marker = AnchorMarker(
        col=col_idx0, colOff=pixels_to_EMU(x_off_px),
        row=row_idx0, rowOff=pixels_to_EMU(y_off_px)
    )
    img.anchor = OneCellAnchor(
        _from=marker,
        ext=XDRPositiveSize2D(pixels_to_EMU(new_w), pixels_to_EMU(new_h))
    )
    ws.add_image(img)


def norm_header(x) -> str:
    """Normalize header text."""
    return str(x).replace("\n", " ").replace("\xa0", " ").strip() if x is not None else ""


def get_po_col_map(ws, header_row=HEADER_ROW):
    """Create header->col index mapping from template."""
    col_map = {}
    for c in range(1, ws.max_column + 1):
        v = norm_header(ws.cell(header_row, c).value)
        if v:
            col_map[v] = c
    return col_map


def copy_column_widths(src_ws, dst_ws):
    """Copy column widths."""
    for col_letter, dim in src_ws.column_dimensions.items():
        if dim.width is not None:
            dst_ws.column_dimensions[col_letter].width = dim.width


def copy_row_style(ws, src_row: int, dst_row: int, max_col: int):
    """Copy style from src_row to dst_row."""
    for c in range(1, max_col + 1):
        src = ws.cell(src_row, c)
        dst = ws.cell(dst_row, c)
        if src.has_style:
            dst._style = src._style


def copy_row_height(ws, src_row: int, dst_row: int):
    """Copy row height."""
    ws.row_dimensions[dst_row].height = ws.row_dimensions[src_row].height


def force_bottom_border(ws, row: int, start_col: int, end_col: int):
    """Force thin bottom border on a row range."""
    thin = Side(style="thin")
    for c in range(start_col, end_col + 1):
        cell = ws.cell(row, c)
        b = _copy(cell.border) if cell.border else Border()
        b.bottom = thin
        cell.border = b


def copy_template_rows(src_ws, dst_ws, src_start, src_end, dst_start, max_col):
    """Copy a row block including merges/styles."""
    row_offset = dst_start - src_start

    for r in range(src_start, src_end + 1):
        for c in range(1, max_col + 1):
            src = src_ws.cell(r, c)
            dst = dst_ws.cell(r + row_offset, c)
            dst.value = src.value
            if src.has_style:
                dst._style = src._style

    for r in range(src_start, src_end + 1):
        dst_ws.row_dimensions[r + row_offset].height = src_ws.row_dimensions[r].height

    for merged in src_ws.merged_cells.ranges:
        if merged.min_row >= src_start and merged.max_row <= src_end and merged.max_col <= max_col:
            dst_ws.merge_cells(
                start_row=merged.min_row + row_offset,
                start_column=merged.min_col,
                end_row=merged.max_row + row_offset,
                end_column=merged.max_col,
            )


def paste_total_block_and_fix(
    ws,
    template_ws,
    total_block_start: int,
    po_last_col: int,
    item_start_row: int,
    last_item_row: int,
    col_amt: str,
    rate_thb_per_cny: float,
):
    """
    Paste total section and update SUM formulas to match new last_item_row.
    """
    copy_template_rows(
        template_ws, ws,
        TEMPLATE_TOTAL_START_ROW, TEMPLATE_TOTAL_END_ROW,
        total_block_start,
        max_col=po_last_col
    )

    r1 = total_block_start
    r_mid = r1 + 1
    r2 = r1 + 2

    sum_pat = re.compile(
        rf'(SUM\()(\$?[A-Z]{{1,3}}\$?){item_start_row}:(\$?[A-Z]{{1,3}}\$?)13(\))',
        flags=re.IGNORECASE
    )
    for row in ws.iter_rows(min_row=r1, max_row=r2, min_col=1, max_col=po_last_col):
        for cell in row:
            v = cell.value
            if isinstance(v, str) and v.startswith("=") and "SUM" in v.upper():
                cell.value = sum_pat.sub(
                    rf'\g<1>\g<2>{item_start_row}:\g<3>{last_item_row}\g<4>',
                    v
                )

    ws[f"O{r1}"].value = None
    ws[f"O{r2}"].value = None

    ws[f"N{r1}"].value = f"=SUM({col_amt}{item_start_row}:{col_amt}{last_item_row})"
    ws[f"N{r_mid}"].value = rate_thb_per_cny
    ws[f"N{r2}"].value = f"=N{r_mid}*N{r1}"


def find_label_cell(ws, label: str, max_row: int = 60, max_col: int = 30):
    """Find an exact label cell."""
    target = norm_text(label)
    for r in range(1, min(max_row, ws.max_row) + 1):
        for c in range(1, min(max_col, ws.max_column) + 1):
            if norm_text(ws.cell(r, c).value) == target:
                return r, c
    return None


# =========================
# PO GENERATION
# =========================
def generate_po_from_combined(
    combined_df: pd.DataFrame,
    vendor_code: str,
    po_date: Optional[datetime.date],
    rate_thb_per_cny: float,
    template_path: str,
    catalog_path: str,
    vendor_info_path: str,
    min_factor: int,
    max_factor: int,
) -> str:
    """
    Generate PO file from filtered combined_df (below MIN only).
    """
    if po_date is None:
        po_date = datetime.date.today()

    os.makedirs(PO_OUTPUT_FOLDER, exist_ok=True)
    vendor_key = str(vendor_code).strip().upper()
    output_path = os.path.join(PO_OUTPUT_FOLDER, f"PO_{vendor_key}_BELOW_MIN.xlsx")

    vendor_map = load_vendor_map(vendor_info_path)
    supplier_name = vendor_map.get(vendor_key, {}).get("name", "")
    supplier_addr = vendor_map.get(vendor_key, {}).get("address", "")

    catalog_map = {}
    if os.path.exists(catalog_path):
        catalog_map = build_catalog_map(catalog_path, vendor_code=vendor_key)

    wb = openpyxl.load_workbook(template_path)
    template_ws = wb[TEMPLATE_SHEET_NAME]
    ws = wb.copy_worksheet(template_ws)
    ws.title = "PO"

    copy_column_widths(template_ws, ws)

    po_cols = get_po_col_map(ws, header_row=HEADER_ROW)

    def find_one(keys):
        for k in keys:
            if k in po_cols:
                return k
        return None

    min_key_old = find_one(["MIN*4", "MIN * 4", "MIN×4", "MIN x4", "MINX4"])
    max_key_old = find_one(["MAX*7", "MAX * 7", "MAX×7", "MAX x7", "MAXX7"])
    if not min_key_old or not max_key_old:
        raise RuntimeError("Cannot find MIN/MAX header in template.")

    min_col_idx = po_cols[min_key_old]
    max_col_idx = po_cols[max_key_old]

    ws.cell(HEADER_ROW, min_col_idx).value = f"MIN*{int(min_factor)}"
    ws.cell(HEADER_ROW, max_col_idx).value = f"MAX*{int(max_factor)}"

    col_min = get_column_letter(min_col_idx)
    col_max = get_column_letter(max_col_idx)

    PO_LAST_COL = max(po_cols.values())

    col_cart = get_column_letter(po_cols["CARTONS"])
    col_tot_order = get_column_letter(po_cols["TOTAL QTY (ORDER)"])
    col_amt = get_column_letter(po_cols["AMOUNT (CNY)"])
    col_thb = get_column_letter(po_cols["THB"])

    ws["H6"] = vendor_key
    ws["H6"].font = Font(color="FF0000", bold=True, size=18)

    pos = find_label_cell(ws, "DATE", max_row=20, max_col=30)
    if pos:
        r, c = pos
        ws.cell(r, c + 1).value = po_date

    pos = find_label_cell(ws, "SUPPLIER", max_row=20, max_col=30)
    if pos and supplier_name:
        r, c = pos
        ws.cell(r, c + 1).value = supplier_name

    pos = find_label_cell(ws, "ADDRESS", max_row=25, max_col=30)
    if pos and supplier_addr:
        r, c = pos
        ws.cell(r, c + 1).value = supplier_addr

    combined_df = combined_df.sort_values(by=["รหัสสินค้า", "รายละเอียดสินค้า"], ascending=[True, True]).reset_index(drop=True)

    current_row = ITEM_START_ROW

    for _, row in combined_df.iterrows():
        line = current_row
        current_row += 1

        copy_row_style(ws, TEMPLATE_ITEM_ROW, line, PO_LAST_COL)
        copy_row_height(ws, TEMPLATE_ITEM_ROW, line)

        buyer_item = str(row["รหัสสินค้า"]).strip()
        cat = catalog_map.get(buyer_item, {})

        qty_per_carton = cat.get("qty_per_carton", "")
        try:
            qty_per_carton_num = float(qty_per_carton) if qty_per_carton not in [None, ""] else 0.0
        except Exception:
            qty_per_carton_num = 0.0

        use_month = int(row["USE_MONTH"]) if not pd.isna(row["USE_MONTH"]) else 0

        yuan = row["หยวน"] if not pd.isna(row["หยวน"]) else None
        try:
            yuan_num = float(yuan) if yuan is not None else None
        except Exception:
            yuan_num = None

        ws.cell(line, po_cols["BUYER ITEM NO."]).value = buyer_item

        if cat.get("img_bytes"):
            add_image_to_cell(ws, f"B{line}", cat["img_bytes"])

        ws.cell(line, po_cols["GOODS DESCRIPTION"]).value = cat.get("goods_desc", row.get("รายละเอียดสินค้า", ""))
        ws.cell(line, po_cols["BRAND"]).value = cat.get("brand", "")
        ws.cell(line, po_cols["MATERIAL"]).value = cat.get("material", "")
        ws.cell(line, po_cols["Weight"]).value = cat.get("weight", "")
        ws.cell(line, po_cols["QTY PER CARTON"]).value = qty_per_carton_num if qty_per_carton_num > 0 else None

        ws.cell(line, po_cols["STOCK GREEN"]).value = float(row["STOCK_GREEN"])
        ws.cell(line, po_cols["STOCK ASIA"]).value = float(row["STOCK_ASIA"])
        ws.cell(line, po_cols["ON ORDER"]).value = float(row["ON_ORDER_TOTAL"])
        ws.cell(line, po_cols["USE MONTH"]).value = use_month

        col_use = get_column_letter(po_cols["USE MONTH"])
        col_sg = get_column_letter(po_cols["STOCK GREEN"])
        col_sa = get_column_letter(po_cols["STOCK ASIA"])
        col_on = get_column_letter(po_cols["ON ORDER"])
        col_tq = get_column_letter(po_cols["TOTAL QTY"])
        col_zan = get_column_letter(po_cols["จน./USE MONTH"])
        col_remain0 = get_column_letter(po_cols["คงเหลือ (จน./USE MONTH เดิม)"])
        col_qpc = get_column_letter(po_cols["QTY PER CARTON"])
        col_green = get_column_letter(po_cols["GREEN"])
        col_asia = get_column_letter(po_cols["ASIA"])
        col_fob = get_column_letter(po_cols["FOB PRICE (CNY)"])

        ws[f"{col_min}{line}"] = f"={col_use}{line}*{int(min_factor)}"
        ws[f"{col_max}{line}"] = f"={col_use}{line}*{int(max_factor)}"

        ws[f"{col_tq}{line}"] = f"={col_sg}{line}+{col_sa}{line}+{col_on}{line}"
        ws[f"{col_zan}{line}"] = f"=ROUND(({col_tq}{line}+{col_tot_order}{line})/{col_use}{line},0)"
        ws[f"{col_remain0}{line}"] = f"=ROUND({col_tq}{line}/{col_use}{line},0)"

        ws[f"{col_cart}{line}"] = f"=ROUND(({col_max}{line}-{col_tq}{line})/{col_qpc}{line},0)"
        ws[f"{col_green}{line}"] = f"={col_cart}{line}*{col_qpc}{line}"
        ws[f"{col_asia}{line}"] = 0
        ws[f"{col_tot_order}{line}"] = f"={col_green}{line}+{col_asia}{line}"

        if yuan_num is not None:
            ws[f"{col_fob}{line}"] = yuan_num
            ws[f"{col_thb}{line}"] = yuan_num * float(rate_thb_per_cny)
        else:
            ws[f"{col_fob}{line}"] = None
            ws[f"{col_thb}{line}"] = None

        ws[f"{col_amt}{line}"] = f"={col_fob}{line}*{col_tot_order}{line}"

    if len(combined_df) > 0:
        last_item_row = ITEM_START_ROW + len(combined_df) - 1
        force_bottom_border(ws, last_item_row, 1, PO_LAST_COL)

        total_block_start = last_item_row + 1
        paste_total_block_and_fix(
            ws=ws,
            template_ws=template_ws,
            total_block_start=total_block_start,
            po_last_col=PO_LAST_COL,
            item_start_row=ITEM_START_ROW,
            last_item_row=last_item_row,
            col_amt=col_amt,
            rate_thb_per_cny=float(rate_thb_per_cny),
        )

    wb.remove(template_ws)
    wb.save(output_path)
    return output_path


def export_vendor_all_items_excel(vendor_rows_all: pd.DataFrame, vendor_code: str, out_folder: str = PO_OUTPUT_FOLDER) -> str:
    """
    Export ALL items for vendor. Highlight rows where TOTAL_QTY_NUM < MIN_NUM.
    """
    os.makedirs(out_folder, exist_ok=True)
    vendor_code = str(vendor_code).strip().upper()
    out_path = os.path.join(out_folder, f"PO_{vendor_code}_ALL_ITEMS.xlsx")

    cols_wanted = [
        "buyer", "รหัสสินค้า", "รายละเอียดสินค้า", "barcode",
        "ยอดขาย_ASIA", "STOCK_ASIA", "ON_ORDER_ASIA", "หยวน_ASIA",
        "ยอดขาย_GREEN", "STOCK_GREEN", "ON_ORDER_GREEN", "หยวน_GREEN",
        "ยอดขาย_TOTAL", "ON_ORDER_TOTAL",
        "USE_MONTH", "TOTAL_QTY_NUM", "MIN_NUM", "MAX_NUM", "หยวน",
    ]
    cols = [c for c in cols_wanted if c in vendor_rows_all.columns]

    df_out = vendor_rows_all.copy()
    if "รหัสสินค้า" in df_out.columns and "รายละเอียดสินค้า" in df_out.columns:
        df_out = df_out.sort_values(["รหัสสินค้า", "รายละเอียดสินค้า"], ascending=[True, True])

    with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
        df_out[cols].to_excel(writer, sheet_name="all_items", index=False)

    wb = openpyxl.load_workbook(out_path)
    ws = wb["all_items"]

    header = {str(ws.cell(1, c).value).strip(): c for c in range(1, ws.max_column + 1)}
    if "TOTAL_QTY_NUM" not in header or "MIN_NUM" not in header:
        wb.save(out_path)
        return out_path

    col_total = header["TOTAL_QTY_NUM"]
    col_min = header["MIN_NUM"]

    for r in range(2, ws.max_row + 1):
        tv = ws.cell(r, col_total).value
        mv = ws.cell(r, col_min).value
        try:
            t = float(tv) if tv is not None else None
            m = float(mv) if mv is not None else None
        except Exception:
            continue
        if t is not None and m is not None and t < m:
            for c in range(1, ws.max_column + 1):
                ws.cell(r, c).fill = HIGHLIGHT_BELOW_MIN

    wb.save(out_path)
    return out_path


def generate_po_streamlit(
    express_asia_path: str,
    express_green_path: str,
    catalog_path: str,
    vendor_info_path: str,
    template_path: str,
    vendor_code: str,
    po_date,
    rate_thb_per_cny: float,
    min_factor: int,
    max_factor: int,
) -> dict:
    """
    Streamlit entry:
      - parse files
      - build combined_all (no vendor filter)
      - export ALL items file for vendor
      - build filtered rows below MIN and create PO if any
    """
    vendor_code = str(vendor_code).strip().upper()

    df_asia, info_asia = parse_express_file(express_asia_path, "ASIA")
    df_green, info_green = parse_express_file(express_green_path, "GREEN")

    months = 1
    if info_asia and info_asia.get("months", 0) > 0:
        months = int(info_asia["months"])
    elif info_green and info_green.get("months", 0) > 0:
        months = int(info_green["months"])

    combined_all = build_combined_all(df_asia, df_green, months=months, min_factor=min_factor, max_factor=max_factor)

    vendor_rows_all = combined_all[combined_all["buyer"] == vendor_code].copy()
    if vendor_rows_all.empty:
        buyers = sorted(set(combined_all["buyer"].dropna().astype(str).tolist()))
        preview = ", ".join(buyers[:40])
        raise ValueError(f"Vendor '{vendor_code}' not found. Parsed buyers (first 40): {preview}")

    path_all = export_vendor_all_items_excel(vendor_rows_all, vendor_code=vendor_code)

    vendor_rows_filtered = vendor_rows_all[vendor_rows_all["TOTAL_QTY_NUM"] < vendor_rows_all["MIN_NUM"]].copy()

    path_filtered = None
    if not vendor_rows_filtered.empty:
        path_filtered = generate_po_from_combined(
            combined_df=vendor_rows_filtered,
            vendor_code=vendor_code,
            po_date=po_date,
            rate_thb_per_cny=float(rate_thb_per_cny),
            template_path=template_path,
            catalog_path=catalog_path,
            vendor_info_path=vendor_info_path,
            min_factor=int(min_factor),
            max_factor=int(max_factor),
        )

    return {
        "po_filtered": path_filtered,
        "po_all_items": path_all,
        "count_all": int(len(vendor_rows_all)),
        "count_filtered": int(len(vendor_rows_filtered)),
    }
